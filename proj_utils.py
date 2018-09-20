import cv2
import os, struct
import matplotlib as plt
from array import array as pyarray
import pandas as pd
from openpyxl import load_workbook
from pylab import *
import numpy as np

def load_NMNIST(dataset="training", digits=range(10), path=r'E:\Users\Shashi\OneDrive\Datasets\Shapes'):
    import os
    import numpy as np
    from pylab import *
    if dataset == "training":
        fname_img = os.path.join(path, 'Shapes_1_1_Train_Features.dat')
        fname_lbl = os.path.join(path, 'Shapes_1_1_Train_Labels.dat')
    elif dataset == "testing":
        fname_img = os.path.join(path, 'Shapes_1_1_Test_Features.dat')
        fname_lbl = os.path.join(path, 'Shapes_1_1_Test_Labels.dat')
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    flbl = open(fname_lbl, 'rb')
    lbl = np.fromfile(flbl, dtype=np.uint8)
    flbl.close()

    fimg = open(fname_img, 'rb')
    img = np.fromfile(fimg, dtype=np.uint8)
    fimg.close()

    size=len(lbl)

    ind = [ k for k in range(size) if lbl[k] in digits ]
    N = len(ind)

    rows=28;cols=28;

    images = zeros((N, rows, cols), dtype=uint8)
    labels = zeros((N, 1), dtype=int8)
    for i in range(len(ind)):
        images[i] = array(img[ ind[i]*rows*cols : (ind[i]+1)*rows*cols ]).reshape((rows, cols))
        labels[i] = lbl[ind[i]]
    return images, labels

def vectortoimg(v,show=True):
    import matplotlib as plt
    import cv2
    import os, struct
    import matplotlib as plt
    from array import array as pyarray
    import pandas as pd
    from openpyxl import load_workbook
    from pylab import *
    import numpy as np
    plt.imshow(v.reshape(28, 28),interpolation='None', cmap='gray')
    plt.axis('off')
    if show:
        plt.show()
        
def vectortoimg_3(v,show=False):
    import matplotlib as plt
    import cv2
    import os, struct
    import matplotlib as plt
    from array import array as pyarray
    import pandas as pd
    from openpyxl import load_workbook
    from pylab import *
    import numpy as np
    plt.imshow(v.reshape(28, 28, 3),interpolation='None')
    plt.axis('off')
    if show:
        plt.show()

def to_rgb5(im):
    im.resize((im.shape[0], im.shape[1], 1))
    return np.repeat(im.astype(np.uint8), 3, 2)

def resize_data(data):
    import cv2
    data_upscaled = np.zeros((data.shape[0], 140, 140, 3))
    for i, img in enumerate(data):
        large_img = cv2.resize(img, dsize=(140, 140), interpolation=cv2.INTER_CUBIC)
        data_upscaled[i] = large_img
    return data_upscaled

def labeled_cluster_accuracy(y_true, y_pred):
    """
    Calculate clustering accuracy. Require scikit-learn installed
    # Arguments
        y: true labels, numpy.array with shape `(n_samples,)`
        y_pred: predicted labels, numpy.array with shape `(n_samples,)`
    # Return
        accuracy, in [0,1]
    """
    y_true = y_true.astype(np.int64)
    assert y_pred.size == y_true.size
    D = max(y_pred.max(), y_true.max()) + 1
    #print(D)
    w = np.zeros((D, D), dtype=np.int64)
    for i in range(y_pred.size):
        w[y_pred[i], y_true[i]] += 1
    #print(w)
    from sklearn.utils.linear_assignment_ import linear_assignment
    ind = linear_assignment(w.max() - w)
    #print([w[i, j] for i, j in ind])
    return sum([w[i, j] for i, j in ind]) * 1.0 / y_pred.size